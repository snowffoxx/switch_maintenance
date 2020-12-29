from datetime import date
from openpyxl import *
from openpyxl.styles import Font, Alignment
from gather import GatherData
from parsing import CiscoParse


# connect device and get device check data.
class DeviceCheck:
    def __init__(self, device):
        self.device = device

    def check_result(self):
        result = dict()
        data = str()
        if self.device['protocol'].lower() == 'telnet':
            if self.device['check']:
                data = GatherData(self.device).gather_telnet()
        elif self.device['protocol'].lower() == 'ssh':
            if self.device['check']:
                data = GatherData(self.device).gather_ssh()
        else:
            print('Not supported protocol: %s - %s' % (self.device['ip'], self.device['protocol'],))

        if data:
            if self.device['vendor'].lower() == 'cisco':
                result = {
                    'ip': self.device['ip'],
                    'hostname': CiscoParse(data).hostname(),
                    'dev_model': CiscoParse(data).dev_model(),
                    'os_version': CiscoParse(data).os_ver(),
                    'uptime': CiscoParse(data).uptime(),
                    'cpu_free': CiscoParse(data).cpu_usage(),
                    'mem_free': CiscoParse(data).mem_usage(),
                    'fan': CiscoParse(data).fan(),
                    'temperature': CiscoParse(data).temperature(),
                    'power': CiscoParse(data).power_supply(),
                }
            if self.device['vendor'].lower() == 'extreme':
                result = {
                    'ip': 'unknown',
                    'hostname': 'unknown',
                    'dev_model': 'unknown',
                    'os_version': 'unknown',
                    'uptime': 'unknown',
                    'cpu_free': 'unknown',
                    'mem_free': 'unknown',
                    'fan': 'unknown',
                    'temperature': 'unknown',
                    'power': 'unknown',
                }
        else:
            print('Not supported vendors: %s - %s' % (self.device['ip'], self.device['vendor'],))
        return result


# Get device information from excel file.
def get_device_info(file_name):
    # file_name = 'swpm.xlsx'         # name of excel file
    device_list = 'device_lists'    # name of worksheet
    devices = list()
    book = load_workbook(file_name)
    sheet = book[device_list]
    for row in sheet.rows:
        tmp = dict()
        tmp['ip'] = row[0].value
        tmp['user'] = row[1].value
        tmp['password'] = row[2].value
        tmp['protocol'] = row[3].value
        tmp['port'] = row[4].value
        tmp['vendor'] = row[5].value
        tmp['check'] = row[6].value
        devices.append(tmp)
        del tmp
    header = {
        'ip': 'ip', 'user': 'user', 'password': 'password', 'protocol': 'protocol',
        'port': 'port', 'vendor': 'vendors', 'check': 'check',
    }
    if header in devices:
        devices.remove(header)
        return devices
    else:
        print('Please check excel file...')
        exit(0)


def create_worksheet(file_name):
    # fname = 'swpm.xlsx'
    today = date.today()
    book = load_workbook(file_name)
    sheet_title = 'report_%s' % (today,)
    header = ['IP', 'Host Name', 'Model', 'Ver.', 'UP Time', 'CPU(% idle)', 'MEM(% idle)', 'Power', 'Fan', 'Temp.']
    try:
        book.create_sheet(title=sheet_title)
        sheet = book[sheet_title]
        # sheet.merge_cells('A1:J1')
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
        sheet['A1'] = 'Maintenance Report - %s' % (today,)
        sheet['A1'].font = Font(size=14, bold=True, underline='single')
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.append(header)
        book.save(file_name)
        book.close()
    except Exception as e:
        print('Some error occurred. \n  %s' % (e,))
        book.close()
    finally:
        book.close()


def device_check():
    results = list()
    devices = get_device_info('swpm.xlsx')
    for i in devices:
        print('%s 장비를 점검합니다' % (i['ip'],))
        r = DeviceCheck(i).check_result()
        results.append(r)
    return results


def generate_report(file_name, result):
    book = load_workbook(file_name)
    today = date.today()
    book = load_workbook(file_name)
    sheet_title = 'report_%s' % (today,)
    sheet = book[sheet_title]
    for i in result:
        data = [
            i['ip'], i['hostname'], i['dev_model'], i['os_version'], i['uptime'],
            i['cpu_free'], i['mem_free'], i['power'], i['fan'], i['temperature']
        ]
        sheet.append(data)
    book.save(file_name)
    book.close()


if __name__ == '__main__':
    print('####### 장비 점검을 시작합니다. #######')
    check = device_check()
    print('>>> 보고서를 작성합니다. <<<')
    create_worksheet('swpm.xlsx')
    generate_report('swpm.xlsx', check)
    print('점검을 완료하였습니다.')
